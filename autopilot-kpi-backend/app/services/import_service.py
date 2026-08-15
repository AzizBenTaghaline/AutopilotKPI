from datetime import datetime
from io import BytesIO

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook

from app.models.enums import ImportEntityType, ImportStatus, UserRole
from app.models.import_record import Import
from app.models.kpi import Kpi
from app.models.user import User
from app.schemas.event_com import EventComCreate
from app.schemas.import_record import ImportResponse
from app.schemas.kpi_entry import KpiEntryCreate
from app.schemas.sav_reclamation import SavReclamationCreate
from app.schemas.sav_retour import SavRetourCreate
from app.services.audit_log_service import AuditLogService
from app.services.kpi_entry_service import KpiEntryService
from app.services.sav_reclamation_service import SavReclamationService
from app.services.sav_retour_service import SavRetourService


def _to_response(record: Import) -> ImportResponse:
    return ImportResponse(
        id=str(record.id),
        nom_fichier=record.nom_fichier,
        type_source=record.type_source,
        entity_type=record.entity_type,
        statut=record.statut,
        nb_lignes=record.nb_lignes,
        nb_lignes_succes=record.nb_lignes_succes,
        nb_lignes_erreur=record.nb_lignes_erreur,
        erreurs=record.erreurs,
        created_by=record.created_by,
        created_at=record.created_at,
    )


def _parse_excel_rows(content: bytes) -> list[dict]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Fichier Excel invalide ou corrompu")

    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)

    try:
        headers = [str(h).strip().lower() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        raise HTTPException(status_code=400, detail="Le fichier est vide")

    rows = []
    for row in rows_iter:
        if all(cell is None for cell in row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def _parse_date(value) -> datetime:
    if isinstance(value, datetime):
        return value
    return datetime.fromisoformat(str(value).strip())


def _error_message(exc: Exception) -> str:
    if isinstance(exc, HTTPException):
        return str(exc.detail)
    return str(exc)


class ImportService:
    @staticmethod
    async def import_kpi_entries(file: UploadFile, current_user: User) -> ImportResponse:
        content = await file.read()
        rows = _parse_excel_rows(content)

        record = Import(
            nom_fichier=file.filename,
            type_source="xlsx",
            entity_type=ImportEntityType.KPI_ENTRY,
            nb_lignes=len(rows),
            created_by=str(current_user.id),
        )
        await record.insert()

        errors = []
        success_count = 0

        for i, row in enumerate(rows, start=2):  # ligne 2 = 1ère ligne de données
            try:
                code = str(row.get("code_kpi", "")).strip()
                kpi = await Kpi.find_one(Kpi.code == code)
                if not kpi:
                    raise ValueError(f"KPI introuvable avec le code '{code}'")

                comment = row.get("commentaire")
                data = KpiEntryCreate(
                    kpi_id=str(kpi.id),
                    value=float(row["valeur"]),
                    period=str(row["periode"]).strip(),
                    comment=str(comment).strip() if comment else None,
                )
                await KpiEntryService.create_entry(data, current_user)
                success_count += 1
            except Exception as e:
                errors.append(f"Ligne {i}: {_error_message(e)}")

        await ImportService._finalize(record, success_count, errors, current_user, "import_kpi_entries")
        return _to_response(record)

    @staticmethod
    async def import_sav_retours(file: UploadFile, current_user: User) -> ImportResponse:
        content = await file.read()
        rows = _parse_excel_rows(content)

        record = Import(
            nom_fichier=file.filename,
            type_source="xlsx",
            entity_type=ImportEntityType.SAV_RETOUR,
            nb_lignes=len(rows),
            created_by=str(current_user.id),
        )
        await record.insert()

        errors = []
        success_count = 0

        for i, row in enumerate(rows, start=2):
            try:
                data = SavRetourCreate(
                    client=str(row["client"]).strip(),
                    cause=str(row["cause"]).strip(),
                    reparation_origine=str(row["reparation_origine"]).strip(),
                    date_retour=_parse_date(row["date_retour"]),
                )
                await SavRetourService.create(data, current_user)
                success_count += 1
            except Exception as e:
                errors.append(f"Ligne {i}: {_error_message(e)}")

        await ImportService._finalize(record, success_count, errors, current_user, "import_sav_retours")
        return _to_response(record)

    @staticmethod
    async def import_sav_reclamations(file: UploadFile, current_user: User) -> ImportResponse:
        content = await file.read()
        rows = _parse_excel_rows(content)

        record = Import(
            nom_fichier=file.filename,
            type_source="xlsx",
            entity_type=ImportEntityType.SAV_RECLAMATION,
            nb_lignes=len(rows),
            created_by=str(current_user.id),
        )
        await record.insert()

        errors = []
        success_count = 0

        for i, row in enumerate(rows, start=2):
            try:
                data = SavReclamationCreate(
                    client=str(row["client"]).strip(),
                    cause=str(row["cause"]).strip(),
                    date_reclamation=_parse_date(row["date_reclamation"]),
                )
                await SavReclamationService.create(data, current_user)
                success_count += 1
            except Exception as e:
                errors.append(f"Ligne {i}: {_error_message(e)}")

        await ImportService._finalize(record, success_count, errors, current_user, "import_sav_reclamations")
        return _to_response(record)

    @staticmethod
    async def _finalize(record: Import, success_count: int, errors: list[str], current_user: User, action: str) -> None:
        record.nb_lignes_succes = success_count
        record.nb_lignes_erreur = len(errors)
        record.erreurs = errors[:50] 
        record.statut = (
            ImportStatus.SUCCESS if not errors
            else ImportStatus.FAILED if success_count == 0
            else ImportStatus.PARTIAL
        )
        await record.save()

        await AuditLogService.log(
            action=action,
            entity_type="Import",
            entity_id=str(record.id),
            performed_by=current_user,
            details={"fichier": record.nom_fichier, "succes": success_count, "erreurs": len(errors)},
        )

    @staticmethod
    async def list_imports(current_user: User) -> list[ImportResponse]:
        if current_user.role in (UserRole.ADMINISTRATEUR, UserRole.MANAGER):
            query = Import.find_all()
        else:
            query = Import.find(Import.created_by == str(current_user.id))

        records = await query.sort(-Import.created_at).to_list()
        return [_to_response(r) for r in records]